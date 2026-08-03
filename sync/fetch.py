#!/usr/bin/env python3
"""Fetch the MAVIR grid xlsx exports and upsert them into versioned daily JSON.

Runs from GitHub Actions every 30 minutes. Pulls the previous 12h + next 12h
window (Europe/Budapest) for both charts (7678 load, 4401 production) — each
fetched independently so one failing does not block the other — parses each,
normalizes rows, and merges them by shared timestamp into
data/<YYYY-MM-DD>.json preferring newer non-null values.

Stdlib only for I/O; openpyxl for xlsx parsing.
"""
from __future__ import annotations

import io
import json
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TZ = ZoneInfo("Europe/Budapest")
BASE_URL = "https://rtdwweb.mavir.hu/rtdwweb/webuser/chart/{chart_id}/export"
WINDOW = timedelta(hours=12)
MAX_ATTEMPTS = 3  # 1 initial + 2 retries
RETRY_BACKOFF_S = 5
TIMEOUT_S = 30

DATA_DIR = Path(__file__).resolve().parent.parent / "data"

# xlsx column letter -> normalized key. B (gross_certified) and H (net_certified)
# are intentionally skipped in v1 (published with multi-day lag, outside window).
LOAD_COLUMN_MAP = {
    "A": "t",
    "C": "gross_est",
    "D": "net_plan_gen",
    "E": "net_plan_load",
    "F": "gross_actual",
    "G": "gross_plan",
    "I": "net_load",
    "J": "net_actual",
    "K": "net_est",
}

# Production (generation) export chart 4401. E (prod_net_actual) publishes with a
# lag and is null inside the ±12h window; mapped null-safely, not relied upon.
PROD_COLUMN_MAP = {
    "A": "t",
    "B": "prod_gross_plan",
    "C": "prod_gross_actual",
    "D": "prod_net_plan",
    "E": "prod_net_actual",
}

# Two independent charts merged into the same daily file by shared `t`.
CHARTS = ((7678, LOAD_COLUMN_MAP), (4401, PROD_COLUMN_MAP))

# Union of every value key across charts (for schema reference).
VALUE_KEYS = [
    v for cm in (LOAD_COLUMN_MAP, PROD_COLUMN_MAP) for k, v in cm.items() if k != "A"
]


def log(msg: str) -> None:
    print(f"[mavir-sync] {msg}", flush=True)


def window_ms(now: datetime | None = None) -> tuple[int, int]:
    """Return (fromTime, toTime) epoch-ms for now-12h .. now+12h in Budapest.

    `now` is floored to the 15-minute grid so the export's bins land on clean
    quarter-hour boundaries and align across runs (stable timestamps => upsert
    merges instead of accumulating near-duplicate points).
    """
    now = now or datetime.now(TZ)
    now = now.replace(minute=now.minute - now.minute % 15, second=0, microsecond=0)
    from_ms = int((now - WINDOW).timestamp() * 1000)
    to_ms = int((now + WINDOW).timestamp() * 1000)
    return from_ms, to_ms


def build_url(chart_id: int, from_ms: int, to_ms: int) -> str:
    base = BASE_URL.format(chart_id=chart_id)
    return (
        f"{base}?exportType=xlsx&fromTime={from_ms}&toTime={to_ms}"
        f"&periodType=min&period=15"
    )


def fetch(url: str) -> bytes:
    """Fetch the xlsx with retries. Raises RuntimeError after all attempts fail."""
    last_err: str = "unknown"
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "mavir-load-sync"})
            with urllib.request.urlopen(req, timeout=TIMEOUT_S) as resp:
                if resp.status != 200:
                    raise RuntimeError(f"HTTP {resp.status}")
                body = resp.read()
            # Validate it is an xlsx (ZIP magic) before accepting.
            if not body.startswith(b"PK"):
                raise RuntimeError("body is not an xlsx (bad magic)")
            log(f"fetch ok on attempt {attempt} ({len(body)} bytes)")
            return body
        except (urllib.error.URLError, RuntimeError, TimeoutError, OSError) as exc:
            last_err = str(exc)
            log(f"fetch attempt {attempt}/{MAX_ATTEMPTS} failed: {last_err}")
            if attempt < MAX_ATTEMPTS:
                time.sleep(RETRY_BACKOFF_S)
    raise RuntimeError(f"fetch failed after {MAX_ATTEMPTS} attempts: {last_err}")


def _norm_ts(raw: str) -> str:
    """'2026.07.31 01:15:00 +0200' -> ISO-8601 '2026-07-31T01:15:00+02:00'."""
    dt = datetime.strptime(raw.strip(), "%Y.%m.%d %H:%M:%S %z")
    return dt.isoformat()


def parse(xlsx_bytes: bytes, colmap: dict = LOAD_COLUMN_MAP) -> list[dict]:
    """Parse an export into normalized points (ISO ts + `colmap`'s value keys)."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    points: list[dict] = []
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r == 1:
            continue  # header
        cell_by_col = {get_column_letter(i): v for i, v in enumerate(row, start=1)}
        raw_t = cell_by_col.get("A")
        if raw_t in (None, ""):
            continue
        point: dict = {"t": _norm_ts(str(raw_t))}
        for col, key in colmap.items():
            if key == "t":
                continue
            v = cell_by_col.get(col)
            point[key] = None if v in (None, "") else float(v)
        points.append(point)
    wb.close()
    return points


def _local_date(iso_ts: str) -> str:
    return datetime.fromisoformat(iso_ts).astimezone(TZ).date().isoformat()


def _merge_point(existing: dict, fresh: dict) -> bool:
    """Fill fields in `existing` from `fresh`'s non-null values. Never overwrite a
    non-null with null. Key-set agnostic (handles load and prod_* alike).
    Returns True if anything changed."""
    changed = False
    for key, new_val in fresh.items():
        if key == "t" or new_val is None:
            continue
        # Fill nulls and pick up genuinely changed values; never null-out.
        if existing.get(key) != new_val:
            existing[key] = new_val
            changed = True
    return changed


def upsert(points: list[dict]) -> list[Path]:
    """Merge points into data/<date>.json files. Returns changed file paths."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    by_date: dict[str, list[dict]] = {}
    for p in points:
        by_date.setdefault(_local_date(p["t"]), []).append(p)

    changed_files: list[Path] = []
    for date, day_points in by_date.items():
        path = DATA_DIR / f"{date}.json"
        if path.exists():
            doc = json.loads(path.read_text(encoding="utf-8"))
            index = {pt["t"]: pt for pt in doc["points"]}
        else:
            doc = {"date": date, "tz": "Europe/Budapest", "points": []}
            index = {}

        changed = False
        for fresh in day_points:
            cur = index.get(fresh["t"])
            if cur is None:
                index[fresh["t"]] = dict(fresh)
                changed = True
            elif _merge_point(cur, fresh):
                changed = True

        if not changed:
            continue

        doc["points"] = [index[t] for t in sorted(index)]
        doc["updated"] = datetime.now(TZ).isoformat()
        path.write_text(
            json.dumps(doc, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        changed_files.append(path)
    return changed_files


def main() -> int:
    from_ms, to_ms = window_ms()
    log(f"window {from_ms}..{to_ms}")
    all_points: list[dict] = []
    ok_charts = 0
    for chart_id, colmap in CHARTS:
        url = build_url(chart_id, from_ms, to_ms)
        try:
            body = fetch(url)
            points = parse(body, colmap)
        except Exception as exc:  # noqa: BLE001 - per-chart isolation
            log(f"chart {chart_id} FAILED: {exc}")
            continue
        log(f"chart {chart_id}: parsed {len(points)} points")
        all_points.extend(points)
        ok_charts += 1
    if ok_charts == 0:
        log("RUN FAILED: all charts failed")
        return 1
    changed = upsert(all_points)
    if changed:
        log(f"changed {len(changed)} file(s): {', '.join(p.name for p in changed)}")
    else:
        log("no data changes")
    return 0


if __name__ == "__main__":
    sys.exit(main())
